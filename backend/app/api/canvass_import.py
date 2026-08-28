from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from app.db.database import get_db
from app.core.deps import require_role
from app.models.material import Material
from app.models.material_type import MaterialType
from app.models.transaction import Transaction
from app.core.inventory import sync_inventory

router = APIRouter(prefix="/transactions", tags=["transactions"])

_write_auth = require_role(["Admin", "Project Coordinator"])

_COLUMNS = ["Date", "Supplier", "Material Type", "Material", "Brand", "Price"]


@router.get("/canvass-import-template")
def download_canvass_import_template(db: Session = Depends(get_db), _=Depends(_write_auth)):
    wb = Workbook()

    sheet = wb.active
    sheet.title = "Canvass"
    sheet.append(_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.append(["2026-08-27", "ABC Hardware Supply", "Wiring", "2.0mm2 THHN Wire", "Omni", 15.00])
    for col, width in zip("ABCDEF", [14, 28, 18, 28, 16, 12]):
        sheet.column_dimensions[col].width = width

    types_sheet = wb.create_sheet("Valid Material Types")
    types_sheet.append(["Material Type"])
    types_sheet[1][0].font = Font(bold=True)
    type_names = [t.name for t in db.query(MaterialType).filter(MaterialType.archived == False).order_by(MaterialType.name).all()]
    for name in type_names:
        types_sheet.append([name])
    types_sheet.column_dimensions["A"].width = 28

    materials_sheet = wb.create_sheet("Existing Materials")
    materials_sheet.append(["Material Type", "Rating/Size"])
    for cell in materials_sheet[1]:
        cell.font = Font(bold=True)
    materials = db.query(Material).filter(Material.archived == False).order_by(Material.material_type, Material.rating_size).all()
    for m in materials:
        materials_sheet.append([m.material_type or "", m.rating_size])
    materials_sheet.column_dimensions["A"].width = 24
    materials_sheet.column_dimensions["B"].width = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=canvass_import_template.xlsx"},
    )


@router.post("/canvass-import")
def import_canvass(file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(_write_auth)):
    wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    sheet = wb["Canvass"] if "Canvass" in wb.sheetnames else wb.active

    valid_types = {
        t.name.strip().lower(): t.name
        for t in db.query(MaterialType).filter(MaterialType.archived == False).all()
    }
    materials_by_key = {
        (m.material_type or "").strip().lower() + "||" + m.rating_size.strip().lower(): m
        for m in db.query(Material).filter(Material.archived == False).all()
    }

    created = 0
    skipped = []
    touched_material_ids = set()

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # blank row, not reported

        date_raw = row[0] if len(row) > 0 else None
        supplier = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
        material_type = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        material_name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        brand = str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "") else None
        price_raw = row[5] if len(row) > 5 else None

        if not date_raw or not material_type or not material_name or price_raw in (None, ""):
            skipped.append({"row": i, "reason": "Missing required field", "material_type": material_type, "material": material_name})
            continue

        if isinstance(date_raw, datetime):
            transaction_date = date_raw.date()
        else:
            try:
                transaction_date = datetime.strptime(str(date_raw).strip(), "%Y-%m-%d").date()
            except ValueError:
                skipped.append({"row": i, "reason": "Invalid date", "material_type": material_type, "material": material_name})
                continue

        matched_type = valid_types.get(material_type.strip().lower())
        if not matched_type:
            skipped.append({"row": i, "reason": "Unrecognized Material Type", "material_type": material_type, "material": material_name})
            continue

        material = materials_by_key.get(matched_type.strip().lower() + "||" + material_name.strip().lower())
        if not material:
            skipped.append({"row": i, "reason": "Material not found", "material_type": matched_type, "material": material_name})
            continue

        try:
            price = float(price_raw)
            if price < 0:
                raise ValueError
        except (ValueError, TypeError):
            skipped.append({"row": i, "reason": "Invalid price", "material_type": matched_type, "material": material_name})
            continue

        tx = Transaction(
            transaction_type="Canvass",
            transaction_date=transaction_date,
            is_office_expense=True,
            project_id=None,
            project_name="Canvass",
            amount=price,
            supplier=supplier,
            materials=[{
                "material_id": material.id,
                "quantity": 1,
                "unit_cost": price,
                "total_cost": price,
                "brand": brand,
                "material_name": material.rating_size,
                "material_type": matched_type,
                "unit": material.unit,
                "use_fifo": False,
            }],
        )
        db.add(tx)
        touched_material_ids.add(material.id)
        created += 1

    db.commit()

    for mid in touched_material_ids:
        sync_inventory(db, mid)

    return {"created": created, "skipped": skipped}
