from io import BytesIO
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from app.db.database import get_db
from app.core.deps import require_role
from app.models.material import Material
from app.models.material_type import MaterialType

router = APIRouter(prefix="/materials", tags=["materials"])

_write_auth = require_role(["Admin"])

_COLUMNS = ["Material Type", "Rating/Size", "Unit", "Description", "Min Stock", "Max Stock"]


@router.get("/import-template")
def download_import_template(db: Session = Depends(get_db), _=Depends(_write_auth)):
    wb = Workbook()

    sheet = wb.active
    sheet.title = "Materials"
    sheet.append(_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.append(["Wiring", "2.0mm2 THHN Wire", "meters", "Single conductor stranded copper wire", 20, 200])
    for col, width in zip("ABCDEF", [18, 28, 12, 40, 12, 12]):
        sheet.column_dimensions[col].width = width

    types_sheet = wb.create_sheet("Valid Material Types")
    types_sheet.append(["Material Type"])
    types_sheet[1][0].font = Font(bold=True)
    names = [t.name for t in db.query(MaterialType).filter(MaterialType.archived == False).order_by(MaterialType.name).all()]
    for name in names:
        types_sheet.append([name])
    types_sheet.column_dimensions["A"].width = 28

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materials_import_template.xlsx"},
    )


@router.post("/import")
def import_materials(file: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(_write_auth)):
    wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    sheet = wb["Materials"] if "Materials" in wb.sheetnames else wb.active

    valid_types = {
        t.name.strip().lower(): t.name
        for t in db.query(MaterialType).filter(MaterialType.archived == False).all()
    }
    existing = {
        (m.material_type or "").strip().lower() + "||" + m.rating_size.strip().lower()
        for m in db.query(Material).filter(Material.archived == False).all()
    }

    created = 0
    skipped = []
    to_add = []

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # blank row, not reported

        material_type = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        rating_size = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        unit = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        description = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else None
        min_stock_raw = row[4] if len(row) > 4 else None
        max_stock_raw = row[5] if len(row) > 5 else None

        if not material_type or not rating_size or not unit:
            skipped.append({"row": i, "reason": "Missing required field", "material_type": material_type, "rating_size": rating_size})
            continue

        matched_type = valid_types.get(material_type.strip().lower())
        if not matched_type:
            skipped.append({"row": i, "reason": "Unrecognized Material Type", "material_type": material_type, "rating_size": rating_size})
            continue

        dedupe_key = matched_type.strip().lower() + "||" + rating_size.strip().lower()
        if dedupe_key in existing:
            skipped.append({"row": i, "reason": "Already exists", "material_type": matched_type, "rating_size": rating_size})
            continue

        try:
            min_stock = int(min_stock_raw) if min_stock_raw not in (None, "") else 0
        except (ValueError, TypeError):
            min_stock = 0
        try:
            max_stock = int(max_stock_raw) if max_stock_raw not in (None, "") else None
        except (ValueError, TypeError):
            max_stock = None

        to_add.append(Material(
            rating_size=rating_size,
            material_type=matched_type,
            unit=unit,
            description=description,
            min_stock=min_stock,
            max_stock=max_stock,
        ))
        existing.add(dedupe_key)  # guard against duplicate rows within the same file
        created += 1

    for m in to_add:
        db.add(m)
    db.commit()

    return {"created": created, "skipped": skipped}
